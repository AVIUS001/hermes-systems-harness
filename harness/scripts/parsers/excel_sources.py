from __future__ import annotations

from pathlib import Path

import openpyxl

from lib.graph import GraphBuilder
from lib.normalize import normalize, slug
from lib.sources import collaborations_path, foundation_id, regulation_paths

CHECKLIST_SHEETS = {"Due Diligence Checklist", "Due Diligence Checklist (2)"}


def ingest_excel_sources(g: GraphBuilder) -> None:
    for reg_id, path in regulation_paths():
        if not path.exists():
            continue
        if "dd" in reg_id or "eval" in reg_id:
            _ingest_dd_workbook(g, path, f"source-{reg_id}", "org-aerial-industries-sg", "dd")
        elif "rrl" in reg_id:
            _ingest_rrl(g, path)
    sybel_dir = collaborations_path()
    if sybel_dir and sybel_dir.exists():
        _ingest_sybel(g, sybel_dir)


def _ingest_dd_workbook(
    g: GraphBuilder,
    path: Path,
    source_id: str,
    owner: str,
    prefix: str,
) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    for sheet_name in wb.sheetnames:
        if sheet_name not in CHECKLIST_SHEETS:
            continue
        ws = wb[sheet_name]
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or len(row) < 4:
                continue
            req_no = normalize(str(row[2] or ""))
            desc = normalize(str(row[3] or ""))
            query = normalize(str(row[4] or "")) if len(row) > 4 else ""
            if not req_no or req_no in ("-", "UAV Requirement-No"):
                continue
            if not desc or desc in ("-", "Description"):
                continue
            nid = f"req-{prefix}-{slug(req_no)}"
            g.upsert({
                "id": nid,
                "type": "requirement",
                "title": f"{req_no}: {desc}",
                "status": "active",
                "owner": "certification",
                "source_path": str(path),
                "metadata": {
                    "requirement_no": req_no,
                    "description": desc,
                    "query": query,
                    "sheet": sheet_name,
                    "row": row_idx,
                    "requirement_class": "in_house_due_diligence",
                },
            })
            g.edge(nid, foundation_id(), "derives_from", "In-house due diligence requirement")
            g.edge(nid, source_id, "hosted_on", f"Defined in {path.name} / {sheet_name}")
            g.edge(nid, owner, "owned_by", "Due diligence requirement owner")
            g.edge(nid, "platform-droneconduct", "satisfies", "Feeds droneCONDUCT due diligence module")
            g.bump(f"{prefix}_requirements", 1)
    wb.close()


def _ingest_rrl(g: GraphBuilder, path: Path) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheets = [
        ("RRL 1-6 Type_Certificate", "tc", "RRL 1-6 Type Certificate"),
        ("RRL 7-9 Operational_Approval", "oa", "RRL 7-9 Operational Approval"),
    ]
    q_index = 0
    for sheet_name, phase_code, phase_label in sheets:
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        current_step = ""
        current_phase = ""
        for row in ws.iter_rows(values_only=True):
            if not row:
                continue
            cols = list(row) + [None] * (8 - len(row))
            # Phase header rows
            if cols[0] and str(cols[0]).strip().isdigit() and cols[1] and not cols[2]:
                current_phase = normalize(str(cols[1]))
                continue
            if cols[2] and str(cols[2]).strip() and not cols[3]:
                current_step = normalize(str(cols[2]))
                continue
            note = cols[3]
            weight = cols[5]
            status = cols[6] if len(cols) > 6 else None
            if not note or not str(note).strip():
                continue
            if str(note).strip() in ("Notes", "Scoring"):
                continue
            try:
                float(weight)
            except (TypeError, ValueError):
                if str(weight) not in ("1", "2", "3", "Yes", "No"):
                    continue
            q_index += 1
            nid = f"req-rrl-{phase_code}-{q_index:03d}"
            title = normalize(str(note))[:120]
            g.upsert({
                "id": nid,
                "type": "requirement",
                "title": title,
                "status": "review_required" if str(status).strip().lower() == "no" else "active",
                "owner": "certification",
                "source_path": str(path),
                "metadata": {
                    "rrl_phase": phase_label,
                    "rrl_phase_code": phase_code,
                    "step": current_step,
                    "phase_section": current_phase,
                    "weight": str(weight),
                    "rrl_status": normalize(str(status or "")),
                    "requirement_class": "regulatory_readiness_level",
                },
            })
            g.edge(nid, foundation_id(), "derives_from", "RRL calculator foundational gate")
            g.edge(nid, "source-rrl-calculator", "hosted_on", f"{sheet_name} row")
            g.edge(nid, "org-aerial-industries-sg", "owned_by", "Certification readiness")
            g.edge(nid, "platform-droneconduct", "verifies", "RRL question in droneCONDUCT bundle")
            reg_id = f"regulation-rrl-{phase_code}"
            g.upsert({
                "id": reg_id,
                "type": "regulation",
                "title": phase_label,
                "status": "active",
                "source_path": str(path),
                "metadata": {"rrl_phase": phase_code},
            })
            g.edge(nid, reg_id, "satisfies", f"RRL gate for {phase_label}")
            g.bump("rrl_requirements", 1)
    wb.close()


def _ingest_sybel(g: GraphBuilder, directory: Path) -> None:
    for xlsx in sorted(directory.glob("*.xlsx")):
        aid = f"artifact-sybel-{slug(xlsx.stem)}"
        g.upsert({
            "id": aid,
            "type": "artifact",
            "title": normalize(xlsx.stem),
            "status": "active",
            "source_path": str(xlsx),
            "metadata": {"format": "xlsx", "folder": "aRiO-COBRA_excel-sheets"},
        })
        g.edge(aid, "artifact-ario-cobra-sheets", "hosted_on", "Sybel collaboration workbook")
        g.edge(aid, "org-sybel-investments", "owned_by", "Sybel Investments collaboration data")
        g.edge(aid, foundation_id(), "derives_from", "Regional SPaaS and investment diligence context")
        g.bump("sybel_artifacts", 1)

        # Parse embedded due diligence checklists (same schema as droneconduct)
        try:
            wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
            for sn in wb.sheetnames:
                if sn in CHECKLIST_SHEETS:
                    wb.close()
                    _ingest_dd_workbook(g, xlsx, aid, "org-sybel-investments", f"sybel-{slug(xlsx.stem)}")
                    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
            wb.close()
        except Exception:
            pass

        # GA_Hub data model rows as interface requirements
        if "GA_Hub" in xlsx.name or "Adalo" in xlsx.name:
            _ingest_ga_hub_fields(g, xlsx)


def _ingest_ga_hub_fields(g: GraphBuilder, path: Path) -> None:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not row or len(row) < 2:
            continue
        field = normalize(str(row[1] or ""))
        if not field or field == "Booking form / Data input item":
            continue
        mandatory = normalize(str(row[2] or "")) if len(row) > 2 else ""
        persona = normalize(str(row[3] or "")) if len(row) > 3 else ""
        nid = f"req-sybel-gahub-{slug(field)}"
        g.upsert({
            "id": nid,
            "type": "requirement",
            "title": f"SPaaS data field: {field}",
            "status": "active",
            "source_path": str(path),
            "metadata": {
                "mandatory": mandatory,
                "persona": persona,
                "row": row_idx,
                "requirement_class": "spaas_data_model",
            },
        })
        g.edge(nid, foundation_id(), "derives_from", "Grower application SPaaS data requirement")
        g.edge(nid, "platform-farmingcourses", "implements", "Training/ops data interoperability")
        g.edge(nid, "org-sybel-investments", "owned_by", "Sybel regional SPaaS")
        g.bump("sybel_gahub_requirements", 1)
    wb.close()
